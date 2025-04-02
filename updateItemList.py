# Author: To'a Heck
# This program is to read through the Excel Spreadsheet
# provided by ***** ******* with the most recent Item List. 

from openpyxl import Workbook, load_workbook
import pyodbc
import json


#Load config file
with open("config.json", "r") as config_file:
    config = json.load(config_file)

#open input file
newTable = load_workbook("newTable.xlsx")
newWs = newTable.active

#initialize list
newItemList = []

#populate "newItemList" list with newTable.xlsx
for row in range(1, newWs.max_row + 1):
  temp = [newWs['A' + str(row)].value, newWs['B' + str(row)].value]
  newItemList.append(temp)

#Connect to DB
devConnectionString = config["devConnectionString"]
prodConnectionString = config["prodConnectionString"]

#prompt user for input
while True:
  userInput = input("Which table would you like to update? (Enter \"dev\" or \"prod\"):  ")
  if userInput == "dev":
    #connect to the dev server
    connection = pyodbc.connect(devConnectionString)
    break
  elif userInput == "prod":
    #connect to production server
    connection = pyodbc.connect(prodConnectionString)
    break

#create a cursor
cursor = connection.cursor()

#get the current ItemList
currentItemList = "SELECT ItemNum, Description FROM WebApp.ItemList;"
cursor.execute(currentItemList)
oldItemList = cursor.fetchall()

#get ItemListArchive table
archiveQuery = "SELECT ItemNum, Description FROM WebApp.ItemListArchive"
cursor.execute(archiveQuery)
ItemListArchive = cursor.fetchall()

#convert recordsets into Python sets
#NOTE: sets have constant-time lookup which will help with the time complexity of this program.
#      Also, these lines convert each list into tuples since Lists cannot be put into sets.
newItemSet = set(tuple(x) for x in newItemList)
archiveSet = set(tuple(x) for x in ItemListArchive)

#build list of records missing from the new list
tempList = []
for i in oldItemList:
  if tuple(i) not in newItemSet:
    tempList.append(i)




#INSERT the pared-down list into ItemListArchive table--------------------------------------------------------------------------
concatStrArchive = "INSERT INTO WebApp.ItemListArchive (ItemNum, Description, Status) VALUES "
for i in tempList:
  if tuple(i[0]) not in archiveSet:
    #check the current string if it contains an apostrophe
    if '\'' in i[1]:
      #update the string to prevent SQL errors
      i[1] = i[1].replace("'", "''")
    #build the concat string
    concatStrArchive = concatStrArchive + "('" + i[0] + "','" + i[1] +"','Closed'),"
#clean up concatStrArchive to be readable by SQL
concatStrArchive = concatStrArchive[:-1] + ";"

print("The following string will be executed to close Items's and insert them into WebApp.ItemListArchive: \n")
print(concatStrArchive)
cursor.execute(concatStrArchive)
connection.commit()




#DELETE the data from ItemList--------------------------------------------------------------------------------------------------
print("Deleting the entire table: \n")
deleteStr = "DELETE FROM WebApp.ItemList"
cursor.execute(deleteStr)
connection.commit()




#INSERT the values from the new table provided by *******---------------------------------------------------------------------
concatStrInsert = "INSERT INTO WebApp.ItemList (ItemNum, Description, Status) VALUES "
for i in range(len(newItemList)):
  currDescription = newItemList[i][1]
  #check the current string if it contains an apostrophe
  if '\'' in newItemList[i][1]:
    #update the string to prevent SQL errors
    currDescription = newItemList[i][1].replace("'","''")
  #build the concat string
  concatStrInsert = concatStrInsert + "('" + newItemList[i][0] + "','" + currDescription +"','Open'),"
#clean up concatStrInsert to be readable by SQL
concatStrInsert = concatStrInsert[:-1] + ";"

print("The following string will be executed to insert into WebApp.ItemList: \n")
print(concatStrInsert)
cursor = connection.cursor()
cursor.execute(concatStrInsert)
connection.commit()




#Close connection
cursor.close()
connection.close()